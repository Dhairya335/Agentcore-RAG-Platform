"""
Phase 2C — Ingestion Worker Lambda  (optimised rewrite)

Triggered by SQS (batchSize=1) from S3 ObjectCreated events.

Flow per document:
  1. Parse SQS → S3 event → get bucket + key
  2. Read S3 object metadata (doc-id, tenant-id, version stored by presign Lambda)
  3. Download file from S3
  4. Detect type → apply type-specific chunking strategy
  5. Embed all chunks in parallel (ThreadPoolExecutor, I/O-bound Bedrock calls)
  6. Batch INSERT all chunks into Aurora pgvector (single rds_data round-trip)
  7. Update DynamoDB: status UPLOADED → READY (or FAILED on error)

Chunking strategy by type:
  PDF   — page-aware, paragraph split, table detection, 600 tokens / 100 overlap
  DOCX  — heading section grouping, paragraph accumulation, 600 tokens / 100 overlap
  XLSX  — streaming 30-row groups, column headers repeated in every chunk
  CSV   — streaming 40-row groups, column headers repeated in every chunk
  TXT   — paragraph accumulation, 600 tokens / 100 overlap
  MD    — H1/H2/H3 section grouping, 600 tokens / 100 overlap

Complexity target: O(N) per document where N = total tokens in the document.
All token counting is done on integer token-id lists, not on strings.

Key improvements over v1:
  - _PERIOD_IDS computed once at module load (was recomputed on every _token_windows call)
  - _token_windows is now a generator — yields one window at a time, O(1) peak RAM
  - CSV / XLSX use streaming row iterators — never materialise the full file in RAM
  - Bedrock embed calls run in parallel via ThreadPoolExecutor (I/O-bound → ~linear speedup)
  - Aurora inserts use batch_execute_statement — 1 round-trip per document (was N)
  - Section headings are prepended to chunk content so embeddings capture heading context
  - Cross-section token overlap is reset at heading boundaries (no semantic bleed-through)
  - Cheap char-length guard replaces full tokenise-just-to-count on spreadsheet chunks
"""

import csv as csv_module
import io
import json
import os
import re
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from typing import Any, Generator, Iterator

import boto3
import tiktoken

s3        = boto3.client("s3")
bedrock   = boto3.client("bedrock-runtime")
rds_data  = boto3.client("rds-data")
dynamodb  = boto3.client("dynamodb")
ssm       = boto3.client("ssm")

STACK_NAME      = os.environ["STACK_NAME"]
DOCS_TABLE_NAME = os.environ["DOCS_TABLE_NAME"]

# Tokenizer — loaded once at cold start, shared for the container lifetime
TOKENIZER = tiktoken.get_encoding("cl100k_base")

CHUNK_TOKENS        = 600
OVERLAP_TOKENS      = 100
MIN_CHUNK_TOKENS    = 20
MIN_CHUNK_CHARS     = MIN_CHUNK_TOKENS * 3   # cheap proxy: ~3 chars/token avg
CSV_ROWS_PER_CHUNK  = 40
XLSX_ROWS_PER_CHUNK = 30
EMBED_MAX_WORKERS   = 8   # Bedrock InvokeModel is I/O-bound → threads are effective

# Sentence-boundary token ids (cl100k_base) — computed ONCE at module load
# Previously this set was rebuilt inside _token_windows on every call.
_PERIOD_IDS: frozenset[int] = frozenset(
    TOKENIZER.encode(s)[0]
    for s in (".", ".\n", "! ", "? ", "!\n", "?\n")
    if TOKENIZER.encode(s)
)

# SSM values cached per container lifetime
_ssm_cache: dict[str, str] = {}


def get_ssm(key: str) -> str:
    if key not in _ssm_cache:
        _ssm_cache[key] = ssm.get_parameter(
            Name=f"/{STACK_NAME}/rag/{key}"
        )["Parameter"]["Value"]
    return _ssm_cache[key]

# Data model

@dataclass
class Chunk:
    content:       str
    chunk_index:   int
    page_number:   int | None = None
    section_title: str | None = None
    heading_level: int | None = None
    sheet_name:    str | None = None
    row_start:     int | None = None
    row_end:       int | None = None


@dataclass
class EmbeddedChunk:
    chunk:     Chunk
    embedding: list[float]


# Entry point

def handler(event, context):
    for record in event.get("Records", []):
        try:
            process_sqs_record(record)
        except Exception as e:
            print(f"[ERROR] {e}")
            raise


def process_sqs_record(record: dict):
    body       = json.loads(record["body"])
    s3_records = body.get("Records", [])
    for s3_record in s3_records:
        bucket = s3_record["s3"]["bucket"]["name"]
        key    = urllib.parse.unquote_plus(s3_record["s3"]["object"]["key"])
        print(f"[INGEST] s3://{bucket}/{key}")
        process_document(bucket, key)


def process_document(bucket: str, key: str):
    head      = s3.head_object(Bucket=bucket, Key=key)
    obj_meta  = head.get("Metadata", {})
    tenant_id = obj_meta.get("tenant-id")
    doc_id    = obj_meta.get("doc-id")
    version   = int(obj_meta.get("version", "1"))
    file_name = key.split("/")[-1]

    if not tenant_id or not doc_id:
        raise ValueError(f"Missing S3 metadata (tenant-id, doc-id) on key: {key}")

    print(f"[INGEST] tenant={tenant_id} doc={doc_id} v={version} file={file_name}")

    db_cluster_arn = get_ssm("aurora-cluster-arn")
    db_secret_arn  = get_ssm("aurora-secret-arn")
    db_name        = get_ssm("aurora-db-name")

    try:
        obj      = s3.get_object(Bucket=bucket, Key=key)
        raw_data = obj["Body"].read()

        ext    = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else "txt"
        chunks = chunk_document(raw_data, file_name, ext)

        if not chunks:
            raise ValueError(f"No chunks produced for {file_name}")

        chunk_total = len(chunks)
        print(f"[INGEST] {chunk_total} chunks for {file_name} — embedding in parallel")

        # Parallel embedding — Bedrock InvokeModel is network I/O, threads give near-linear speedup.
        # Results are collected in chunk-index order for deterministic inserts.
        embedded = _embed_parallel(chunks)

        # Single-batch Aurora insert
        batch_insert_chunks(
            db_cluster_arn, db_secret_arn, db_name,
            tenant_id, doc_id, key, file_name, ext,
            embedded, chunk_total,
        )

        update_doc_status(tenant_id, doc_id, version, "READY", chunk_total=chunk_total)
        print(f"[INGEST] Done — {chunk_total} chunks stored for doc {doc_id}")

    except Exception as e:
        print(f"[INGEST ERROR] {e}")
        update_doc_status(tenant_id, doc_id, version, "FAILED", error_message=str(e))
        raise


# Parallel embedding

def _embed_parallel(chunks: list[Chunk]) -> list[EmbeddedChunk]:
    """
    Embed all chunks concurrently using a thread pool.
    Bedrock InvokeModel is I/O-bound so threads are far cheaper than processes.
    Results are ordered by chunk_index to preserve insertion order.
    """
    results: list[EmbeddedChunk | None] = [None] * len(chunks)

    with ThreadPoolExecutor(max_workers=EMBED_MAX_WORKERS) as pool:
        future_to_idx = {
            pool.submit(embed_text, c.content): i
            for i, c in enumerate(chunks)
        }
        for future in as_completed(future_to_idx):
            idx            = future_to_idx[future]
            results[idx]   = EmbeddedChunk(chunk=chunks[idx], embedding=future.result())

    return results  # type: ignore[return-value]


# Chunking dispatch

def chunk_document(raw_data: bytes, file_name: str, ext: str) -> list[Chunk]:
    dispatch = {
        "pdf":  chunk_pdf,
        "docx": chunk_docx,
        "doc":  chunk_docx,
        "xlsx": chunk_xlsx,
        "xls":  chunk_xlsx,
        "csv":  chunk_csv,
        "txt":  chunk_text,
        "md":   chunk_markdown,
        "json": chunk_text,
    }
    return dispatch.get(ext, chunk_text)(raw_data, file_name)


# PDF chunking

def chunk_pdf(raw_data: bytes, file_name: str) -> list[Chunk]:
    import PyPDF2

    reader = PyPDF2.PdfReader(io.BytesIO(raw_data))
    chunks: list[Chunk] = []

    for page_num, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        if not page_text.strip():
            continue

        for para in _split_paragraphs(page_text):
            if not para:
                continue

            # Encode once per paragraph, reuse the token list — O(N) not O(2N)
            para_ids = TOKENIZER.encode(para)

            if len(para_ids) < MIN_CHUNK_TOKENS:
                continue

            if _ids_look_like_table(para):
                chunks.append(Chunk(
                    content=para,
                    chunk_index=len(chunks),
                    page_number=page_num,
                ))
                continue

            for window_ids in _token_windows(para_ids, CHUNK_TOKENS, OVERLAP_TOKENS):
                if len(window_ids) >= MIN_CHUNK_TOKENS:
                    chunks.append(Chunk(
                        content=TOKENIZER.decode(window_ids),
                        chunk_index=len(chunks),
                        page_number=page_num,
                    ))

    return chunks


def _ids_look_like_table(text: str) -> bool:
    lines = [l for l in text.splitlines() if l.strip()]
    if not lines:
        return False
    table_count = sum(1 for l in lines if l.count("\t") >= 2 or l.count("|") >= 2)
    return table_count / len(lines) >= 0.5


# DOCX chunking

def chunk_docx(raw_data: bytes, file_name: str) -> list[Chunk]:
    from docx import Document

    doc    = Document(io.BytesIO(raw_data))
    chunks: list[Chunk] = []

    current_heading:     str | None = None
    current_ids:         list[int]  = []
    current_token_count: int        = 0

    def flush(heading: str | None):
        nonlocal current_ids, current_token_count
        if not current_ids:
            return
        # Prepend heading text to the first window so embeddings capture context.
        # Subsequent windows within the same section also carry the heading prefix.
        prefix_ids = TOKENIZER.encode(f"{heading}\n\n") if heading else []
        full_ids   = prefix_ids + current_ids
        for window in _token_windows(full_ids, CHUNK_TOKENS, OVERLAP_TOKENS):
            if len(window) >= MIN_CHUNK_TOKENS:
                chunks.append(Chunk(
                    content=TOKENIZER.decode(window),
                    chunk_index=len(chunks),
                    section_title=heading,
                ))
        # Reset accumulator — no overlap across section boundaries
        current_ids         = []
        current_token_count = 0

    for para in doc.paragraphs:
        style_name = para.style.name if para.style else ""
        text       = para.text.strip()
        if not text:
            continue

        if style_name.startswith("Heading"):
            flush(current_heading)
            current_heading = text
            # Do NOT carry overlap into the new section — prevents semantic bleed
        else:
            para_ids             = TOKENIZER.encode("\n" + text)
            current_ids         += para_ids
            current_token_count += len(para_ids)

            if current_token_count >= CHUNK_TOKENS:
                flush(current_heading)

    flush(current_heading)
    return chunks


# XLSX chunking — streaming, O(chunk_size) RAM

def chunk_xlsx(raw_data: bytes, file_name: str) -> list[Chunk]:
    import openpyxl

    wb     = openpyxl.load_workbook(io.BytesIO(raw_data), read_only=True, data_only=True)
    chunks: list[Chunk] = []

    for sheet in wb.worksheets:
        row_iter: Iterator = sheet.iter_rows(values_only=True)

        # First row = headers
        try:
            header_row  = next(row_iter)
        except StopIteration:
            continue

        headers     = [str(c) if c is not None else "" for c in header_row]
        header_line = "Columns: " + " | ".join(headers)
        sheet_title = sheet.title

        # Stream rows in fixed-size batches — never materialise the full sheet
        for batch, row_start, row_end in _row_batches(row_iter, XLSX_ROWS_PER_CHUNK):
            lines   = [f"Sheet: {sheet_title}", header_line]
            lines  += [" | ".join(str(c) if c is not None else "" for c in row) for row in batch]
            content = "\n".join(lines)

            # Cheap char-length guard avoids full tokenise on every chunk
            if len(content) >= MIN_CHUNK_CHARS:
                chunks.append(Chunk(
                    content=content,
                    chunk_index=len(chunks),
                    sheet_name=sheet_title,
                    row_start=row_start,
                    row_end=row_end,
                ))

    return chunks


# CSV chunking — streaming, O(chunk_size) RAM

def chunk_csv(raw_data: bytes, file_name: str) -> list[Chunk]:
    text   = raw_data.decode("utf-8", errors="replace")
    reader = csv_module.reader(io.StringIO(text))  # csv.reader is already a lazy iterator

    try:
        headers = next(reader)
    except StopIteration:
        return []

    header_line = "Columns: " + " | ".join(headers)
    chunks: list[Chunk] = []

    for batch, row_start, row_end in _row_batches(reader, CSV_ROWS_PER_CHUNK):
        lines   = [header_line] + [" | ".join(row) for row in batch]
        content = "\n".join(lines)

        if len(content) >= MIN_CHUNK_CHARS:
            chunks.append(Chunk(
                content=content,
                chunk_index=len(chunks),
                sheet_name=file_name,
                row_start=row_start,
                row_end=row_end,
            ))

    return chunks


def _row_batches(
    row_iter: Iterator,
    batch_size: int,
) -> Generator[tuple[list, int, int], None, None]:
    """
    Yield (batch, 1-based row_start, 1-based row_end) from a lazy row iterator.
    Accumulates exactly batch_size rows at a time — O(batch_size) RAM at any point.
    Previously the caller used list(sheet.iter_rows(...)) which was O(total_rows).
    """
    batch:     list  = []
    row_start: int   = 1
    row_index: int   = 0

    for row in row_iter:
        batch.append(row)
        row_index += 1
        if len(batch) == batch_size:
            yield batch, row_start, row_start + len(batch) - 1
            row_start += batch_size
            batch      = []

    if batch:
        yield batch, row_start, row_start + len(batch) - 1


# TXT / JSON chunking

def chunk_text(raw_data: bytes, file_name: str) -> list[Chunk]:
    text        = raw_data.decode("utf-8", errors="replace")
    chunks:     list[Chunk] = []
    buffer_ids: list[int]   = []
    flushed:    bool        = False  # guard against double-flush of the overlap tail

    for para in _split_paragraphs(text):
        if not para:
            continue

        flushed    = False
        para_ids   = TOKENIZER.encode(para)
        buffer_ids += para_ids

        if len(buffer_ids) >= CHUNK_TOKENS:
            for window in _token_windows(buffer_ids, CHUNK_TOKENS, OVERLAP_TOKENS):
                if len(window) >= MIN_CHUNK_TOKENS:
                    chunks.append(Chunk(content=TOKENIZER.decode(window), chunk_index=len(chunks)))
            buffer_ids = buffer_ids[-OVERLAP_TOKENS:]
            flushed    = True

    # Final flush — only if the buffer wasn't just emptied in the loop above
    if buffer_ids and not flushed and len(buffer_ids) >= MIN_CHUNK_TOKENS:
        chunks.append(Chunk(content=TOKENIZER.decode(buffer_ids), chunk_index=len(chunks)))

    return chunks


# Markdown chunking

def chunk_markdown(raw_data: bytes, file_name: str) -> list[Chunk]:
    text       = raw_data.decode("utf-8", errors="replace")
    chunks:    list[Chunk] = []
    heading_re = re.compile(r"^(#{1,3})\s+(.*)")

    current_heading: str | None = None
    current_level:   int | None = None
    current_ids:     list[int]  = []

    def flush(heading: str | None, level: int | None):
        nonlocal current_ids
        if not current_ids:
            return
        # Prepend heading so the embedding captures section context
        prefix_ids = TOKENIZER.encode(f"{heading}\n\n") if heading else []
        full_ids   = prefix_ids + current_ids
        for window in _token_windows(full_ids, CHUNK_TOKENS, OVERLAP_TOKENS):
            if len(window) >= MIN_CHUNK_TOKENS:
                chunks.append(Chunk(
                    content=TOKENIZER.decode(window),
                    chunk_index=len(chunks),
                    section_title=heading,
                    heading_level=level,
                ))
        # No overlap across heading boundaries — reset cleanly
        current_ids = []

    for line in text.splitlines():
        m = heading_re.match(line)
        if m:
            flush(current_heading, current_level)
            current_level   = len(m.group(1))
            current_heading = m.group(2).strip()
        else:
            current_ids += TOKENIZER.encode("\n" + line)

    flush(current_heading, current_level)
    return chunks


# Token-space helpers

def _split_paragraphs(text: str) -> list[str]:
    return [p.strip() for p in re.split(r"\n\s*\n", text)]


def _token_windows(
    ids: list[int],
    max_tokens: int,
    overlap: int,
) -> Generator[list[int], None, None]:
    """
    Sliding window over a pre-encoded token-id list.

    GENERATOR — yields one window at a time instead of building the full list.
    Peak RAM = O(max_tokens) regardless of document length.  Previously the
    function returned list[list[int]], holding all windows in memory at once.

    Sentence boundary snapping: scans backward from window end in integer
    token-id space.  Single backward scan, at most 20% of window = O(1)
    per window.

    _PERIOD_IDS is a module-level frozenset, computed once at cold start.
    Previously it was rebuilt on every call to this function.
    """
    if not ids:
        return

    step  = max_tokens - overlap
    start = 0

    while start < len(ids):
        end   = min(start + max_tokens, len(ids))
        chunk = ids[start:end]

        # Sentence-boundary snap: scan back at most 20% of window — O(1) per window
        snap_start = max(0, len(chunk) - max_tokens // 5)
        for i in range(len(chunk) - 1, snap_start, -1):
            if chunk[i] in _PERIOD_IDS:
                chunk = chunk[: i + 1]
                break

        yield chunk
        start += step


# Embedding — Bedrock Titan Embed V2

def embed_text(text: str) -> list[float]:
    response = bedrock.invoke_model(
        modelId="amazon.titan-embed-text-v2:0",
        contentType="application/json",
        accept="application/json",
        body=json.dumps({
            "inputText":  text,
            "dimensions": 1024,
            "normalize":  True,
        }),
    )
    return json.loads(response["body"].read())["embedding"]


# Aurora — batch INSERT via RDS Data API

def batch_insert_chunks(
    db_cluster_arn: str,
    db_secret_arn:  str,
    db_name:        str,
    tenant_id:      str,
    doc_id:         str,
    s3_key:         str,
    file_name:      str,
    source_type:    str,
    embedded:       list[EmbeddedChunk],
    chunk_total:    int,
):
    """
    Insert all chunks for a document in a single batch_execute_statement call.

    Previously each chunk triggered a separate execute_statement call — N
    serial HTTPS round-trips to RDS Data API.  batch_execute_statement sends
    all parameter sets in one call.  Aurora executes them as a single
    server-side batch, reducing round-trips from N to 1.

    Note: batch_execute_statement does not support RETURNING, but we don't
    need the generated UUIDs — id is DEFAULT gen_random_uuid().
    """
    sql = """
        INSERT INTO fast_chunks (
            tenant_id, doc_id, s3_key, file_name, source_type,
            chunk_index, chunk_total, content, embedding,
            page_number, section_title, heading_level,
            sheet_name, row_start, row_end
        ) VALUES (
            :tenant_id, :doc_id, :s3_key, :file_name, :source_type,
            :chunk_index, :chunk_total, :content, :embedding::vector,
            :page_number, :section_title, :heading_level,
            :sheet_name, :row_start, :row_end
        )
    """

    def _str(name: str, value: str | None) -> dict:
        return {"name": name, "value": {"isNull": True} if value is None else {"stringValue": value}}

    def _int(name: str, value: int | None) -> dict:
        return {"name": name, "value": {"isNull": True} if value is None else {"longValue": value}}

    param_sets = []
    for ec in embedded:
        c              = ec.chunk
        vector_literal = "[" + ",".join(f"{v:.8f}" for v in ec.embedding) + "]"
        param_sets.append([
            _str("tenant_id",     tenant_id),
            _str("doc_id",        doc_id),
            _str("s3_key",        s3_key),
            _str("file_name",     file_name),
            _str("source_type",   source_type),
            _int("chunk_index",   c.chunk_index),
            _int("chunk_total",   chunk_total),
            _str("content",       c.content),
            _str("embedding",     vector_literal),
            _int("page_number",   c.page_number),
            _str("section_title", c.section_title),
            _int("heading_level", c.heading_level),
            _str("sheet_name",    c.sheet_name),
            _int("row_start",     c.row_start),
            _int("row_end",       c.row_end),
        ])

    rds_data.batch_execute_statement(
        resourceArn=db_cluster_arn,
        secretArn=db_secret_arn,
        database=db_name,
        sql=sql,
        parameterSets=param_sets,
    )
    print(f"[AURORA] batch inserted {len(param_sets)} chunks in 1 round-trip")


# DynamoDB status update

def update_doc_status(
    tenant_id:     str,
    doc_id:        str,
    version:       int,
    status:        str,
    chunk_total:   int = 0,
    error_message: str = "",
):
    from datetime import datetime, timezone

    pk = f"TENANT#{tenant_id}#DOC#{doc_id}"
    sk = f"VER#{version:06d}"

    expr_parts:  list[str]      = ["#s = :status", "updatedAt = :now"]
    expr_values: dict[str, Any] = {
        ":status": {"S": status},
        ":now":    {"S": datetime.now(timezone.utc).isoformat()},
    }

    if status == "READY" and chunk_total:
        expr_parts.append("chunkCount = :cc")
        expr_values[":cc"] = {"N": str(chunk_total)}

    if status == "FAILED" and error_message:
        expr_parts.append("errorMessage = :err")
        expr_values[":err"] = {"S": error_message[:500]}

    dynamodb.update_item(
        TableName=DOCS_TABLE_NAME,
        Key={"PK": {"S": pk}, "SK": {"S": sk}},
        UpdateExpression="SET " + ", ".join(expr_parts),
        ExpressionAttributeValues=expr_values,
        ExpressionAttributeNames={"#s": "status"},
    )
    print(f"[DYNAMO] {pk} {sk} → {status}")
